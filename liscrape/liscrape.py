import os, csv, time, logging, traceback, random
import concurrent.futures, queue, threading
import PySimpleGUI as sg
import pandas as pd
import ujson as json

from openpyxl import load_workbook
from linkedin_api import Linkedin


# TODO add automated scraping capability: "scrape first-degree contacts"
# TODO check if profile is duplicate with URL before API call


class History:
	'''
	History class loads, stores, and enforces a simple API call-limit to prevent
	users from getting themselves banned by spamming Linkedin's API.
	'''
	def __init__(self, session):
		self.parent_session = session
		self.call_count = 0
		self.hourly_limit = 90
		self.history = {}


	def load(self):
		'''
		Load history from configuration file
		'''
		if not os.path.isfile('config.json'):
			with open('config.json', 'w') as config_file:
				config = {'users': {}, 'history': {}, 'theme': None}
				json.dump(config, config_file, indent=4)

			return {}

		with open('config.json', 'r') as config_file:
			try:
				config = json.load(config_file)
				self.call_count = len(config['history'].keys())
				return config['history']
			except Exception as error:
				logging.exception(error)
				os.remove('config.json')
				return {}


	def store(self):
		'''
		Store history into configuration file
		'''
		if os.path.isfile('config.json'):
			with open('config.json', 'r') as config_file:
				config = json.load(config_file)

			with open('config.json', 'w') as config_file:
				config['history'] = self.history
				json.dump(config, config_file, indent=4)
		else:
			with open('config.json', 'w') as config_file:
				config = {'users': {}, 'history': {}, 'theme': None}
				config['history'] = self.history
				json.dump(config, config_file, indent=4)


	def add(self, user_id, ignore_duplicates):
		'''
		Add a user profile into history, with the current unix time stamp
		'''
		self.call_count += 1
		not_added = False if user_id in self.history.values() else True
		self.history[time.time()] = user_id

		return not_added if not ignore_duplicates else True


	def check_validity(self):
		'''
		Checks if we have API calls left in our quota, and verifies
		that the current profile is not a duplicate.
		'''
		if self.parent_session.debug:
			return True, None

		if self.call_count > self.hourly_limit:
			filtered_history = {}
			for key, val in self.history.items():
				if time.time() - float(key) <= 3600:
					filtered_history[key] = val

			self.call_count = len(filtered_history.keys())
			if self.call_count > self.hourly_limit:
				earliest_call = min([float(key) for key in self.history.keys()])
				since_call = time.time() - earliest_call
				return False, f'{int((3600 - since_call) / 60)} minutes'

		return True, None


class GUI:
	def __init__(self, session):
		self.parent_session = session
		self.window = None
		self.secondary_window = None


	def display_signin_screen(self):
		layout = [
			[sg.Text('Sign in to Linkedin to continue', font=('Helvetica Bold', 11))],
			[sg.Text('Username (email)', font=('Helvetica', 11), size=(15, None)), sg.InputText(key="username")],
			[sg.Text('Password', font=('Helvetica', 11), size=(15, None)), sg.InputText(key="password")],
			[	
				sg.Text('Select a stored login', size=(15, None), font=('Helvetica', 11)),
				sg.Listbox(
					self.parent_session.load_configuration(), select_mode='LISTBOX_SELECT_MODE_SINGLE', 
					enable_events=True, size=(40, 1 + len(self.parent_session.load_configuration())),
					key='-USERNAME-', no_scrollbar=True
					)
			],
			[
				sg.Button('Sign in', font=('Helvetica', 11)), 
				sg.Checkbox('Remember me', key='remember'),
				sg.Checkbox('Refresh cookies', key='cookies'),
				sg.Checkbox('Debug mode', key='debug_mode'),
				sg.Checkbox('Dark theme' if self.parent_session.load_theme() == 'SystemDefault' else 'Light theme', key='theme_switch', enable_events=True),
			],
				[sg.Output(size=(80, 20), font=('Helvetica', 11), key='output_window')],
				[
					sg.Button('Tools', font=('Helvetica', 11), key='debug_screen'),
					sg.Button('Show log', font=('Helvetica', 11), key='show_log'), 
					sg.Text(f'Log file length: {self.parent_session.get_log_length()} lines', key='log_length' , font=('Helvetica', 11))
				]
		]

		self.window = sg.Window(f'Liscrape version {self.parent_session.version}', layout=layout, resizable=True, grab_anywhere=True)


	def display_sheet_screen(self):
		layout = [
			[sg.Text('Choose file to store contacts in', font=('Helvetica', 11))],
			[sg.FileBrowse(), sg.Input(key="sheet_path")],
			[sg.Text('Supported file types: .xls, .xlsx, .xlsm, .csv', font=('Helvetica', 9))],
			[sg.Button('OK'), sg.Button('Use default')]
		]

		self.window = sg.Window(f'Liscrape version {self.parent_session.version}', layout)


	def display_debug_screen(self):
		layout = [
			[sg.Text('Debug settings and tools', font=('Helvetica', 11))],
			[sg.Text('Clear program log', font=('Helvetica', 11), size=(20, None)), sg.Button('Clear log')],
			[sg.Text('Clear configuration file', font=('Helvetica', 11), size=(20, None)), sg.Button('Clear configuration')],
			[sg.Text('Clear all stored contacts', font=('Helvetica', 11), size=(20, None)), sg.Button('Remove contacts')],
			[sg.Text('Run automated tests', font=('Helvetica', 11), size=(20, None)), sg.Button('Run tests')]
		]

		self.secondary_window = sg.Window(f'Liscrape version {self.parent_session.version}', layout)


	def display_main_screen(self):
		layout = [
			[sg.Text('Signed in as:', font=('Helvetica', 11)), sg.Text(f'{self.parent_session.username}', font=('Helvetica', 11), text_color='Blue')],
			[sg.Text('Contact to store (URL)', font=('Helvetica', 11)), sg.InputText(key="profile_url")],
			[sg.Button('Store contact', font=('Helvetica', 11)), sg.Text(f'{self.parent_session.parsed} contacts stored (this session)\t', key='parsed', font=('Helvetica', 11))],
			[sg.Output(size=(60, 15), font=('Helvetica', 11))],
			[
				sg.Text(f'Contacts in file: {self.parent_session.total_parsed}', font=('Helvetica', 11), key='total_parsed', size=(15, None)), 
				sg.Text(f'Session path: {self.parent_session.sheet_path}', font=('Helvetica', 11))
			]]

		self.window = sg.Window(title=f'Liscrape version {self.parent_session.version}',
			layout=layout, resizable=True, grab_anywhere=True)


class Session:
	def __init__(self):
		self.version = '1.3.1'
		self.username = None
		self.password = None
		self.authenticated = False

		# sheet properties
		self.sheet_path = None
		self.sheet_type = None
		self.default_sheet_type = 'excel'

		# keep track of parse counts in memory
		self.total_parsed = 0
		self.parsed = 0

		# additional options
		self.log_filename = 'liscrape-log.log'
		self.ignore_duplicates = False
		self.debug = False

		# gui
		self.gui = GUI(self)

		# history, load validity
		self.history = History(self)
		self.history.history = self.history.load()
		self.history.check_validity()


	def start_log(self):
		logging.basicConfig(
		filename=self.log_filename, level=logging.DEBUG,
		format='%(asctime)s %(message)s', datefmt='%d/%m/%Y %H:%M:%S'
		)


	def get_log_length(self):
		if not os.path.isfile(self.log_filename):
			return 0

		with open(self.log_filename, 'r') as log_file:
			return sum(1 for row in log_file)


	def load_log(self):
		if self.get_log_length() == 0:
			return '-- Log is empty --\n'

		with open(session.log_filename, 'r') as log_file:
			return log_file.read()


	def clear_log(self):
		try:
			logging.shutdown()
		except Exception as e:
			sg.popup(traceback.format_exc())
			logging.exception(f'Exception attempting to shutdown logging: {e}')
			return

		if os.path.isfile(self.log_filename):
			os.remove(self.log_filename)
			sg.popup(f'Log file {self.log_filename} successfully removed!')

			# restart log, refresh log length
			self.start_log()
			self.gui.window['log_length'].update(f'Log file length: {self.get_log_length()} lines')
			self.gui.window['output_window'].update(self.load_log())
		else:
			sg.popup('Nothing to remove!')


	def remove_contacts(self):
		if os.path.isfile('linkedin_scrape.xlsx'):
			os.remove('linkedin_scrape.xlsx')
			sg.popup('Contacts file linkedin_scrape.xlsx removed!')

		if os.path.isfile('linkedin_scrape.csv'):
			os.remove('linkedin_scrape.csv')
			sg.popup('Contacts file linkedin_scrape.xlsx csv!')


	def clear_config(self):
		self.history.history = {}
		if os.path.isfile('config.json'):
			with open('config.json', 'r') as config_file:
				config = json.load(config_file)
				users = config['users']

			with open('config.json', 'w') as config_file:
				config = {'users': users, 'history': {}, 'theme': None}
				json.dump(config, config_file, indent=4)
				sg.popup('Configuration file cleared!')


	def load_sheet_length(self):
		if not os.path.isfile(self.sheet_path):
			logging.info(f'Sheet {self.sheet_path} does not exist: returning total_parsed=0')
			self.total_parsed = 0
		else:
			logging.info(f'Sheet {self.sheet_path} exists: getting length.')
			if self.sheet_type == 'csv':
				with open(self.sheet_path, 'r') as csv_file:
					csv_reader = csv.reader(csv_file)
					self.total_parsed = sum(1 for row in csv_reader)
			elif self.sheet_type == 'excel':
				df = pd.read_excel(self.sheet_path)
				self.total_parsed = len(df.index)

		return self.total_parsed


	def load_configuration(self):
		if not os.path.isfile('config.json'):
			return []

		with open('config.json', 'r') as config_file:
			try:
				config = json.load(config_file)
			except Exception as error:
				logging.exception(error)
				os.remove('config.json')
				return ()

			return tuple(config['users'].keys()) if len(config['users'].keys()) > 0 else ()


	def load_theme(self):
		if not os.path.isfile('config.json'):
			return 'SystemDefault'

		with open('config.json', 'r') as config_file:
			try:
				config = json.load(config_file)
			except Exception as error:
				logging.exception(error)
				os.remove('config.json')
				return ()

			try:
				return config['theme'] if config['theme'] != None else 'SystemDefault'
			except KeyError:
				with open('config.json', 'r') as config_file:
					config = json.load(config_file)

				config['theme'] = None
				with open('config.json', 'w') as config_file:
					json.dump(config, config_file, indent=4)

				return self.load_theme()


	def save_theme(self, theme):
		if not os.path.isfile('config.json'):
			with open('config.json', 'w') as config_file:
				config = {'users': {}, 'history': {}, 'theme': None}
		else:
			with open('config.json', 'r') as config_file:
				config = json.load(config_file)

		config['theme'] = theme
		with open('config.json', 'w') as config_file:
			json.dump(config, config_file, indent=4)

		return True


	def load_password_from_config(self, username):
		with open('config.json', 'r') as config_file:
			config = json.load(config_file)

		try:
			return config['users'][username]
		except:
			sg.popup('Error finding password from configuration!', title='Error', keep_on_top=True)
			raise Exception('Error finding password from configuration!')


	def store_login(self, username, password):
		if not os.path.isfile('config.json'):
			with open('config.json', 'w') as config_file:
				config = {'users': {}, 'history': {}, 'theme': None}
		else:
			with open('config.json', 'r') as config_file:
				config = json.load(config_file)

		config['users'][username] = password
		with open('config.json', 'w') as config_file:
			json.dump(config, config_file, indent=4)

		return True


	def sign_in(self, username, password, remember_login, refresh_cookies):
		self.username = username
		self.password = password
		auth_success = self.authenticate(refresh_cookies)

		if self.authenticated and remember_login:
			success = self.store_login(username, password)
			if success:
				print('Login stored into config file successfully!')

		return auth_success


	def authenticate(self, refresh_cookies):
		try:
			self.application = Linkedin(self.username, self.password, debug=True, refresh_cookies=refresh_cookies)
			self.authenticated = True
			return True
		except Exception as error:
			logging.exception(error)
			if 'BAD_EMAIL' in error.args:
				sg.popup('Incorrect email: try again.', title='Incorrect email', keep_on_top=True)
			elif 'CHALLENGE' in error.args:
				sg.popup('Error: LinkedIn requires a sign-in challenge.', title='Linkedin error', keep_on_top=True)
			elif 'Expecting value: line 1 column 1 (char 0)' in error.args:
				sg.popup('Linkedin is refusing to sign in. Please try again later.', title='Unable to sign in', keep_on_top=True)
			else:
				sg.popup(f'Error arguments: {error.args}\n{traceback.format_exc()}', title='Unhandled exception', keep_on_top=True)

			return False


	# perform the API calls
	def linkedin_api_call(self, queue, event):
		while not event.is_set() or not queue.empty():
			profile_url = queue.get()
			if not self.debug:
				try:
					# two API requests: profile and contact info
					profile = self.application.get_profile(profile_url)
				except Exception as error:
					logging.exception(f'Error loading profile: {error}')
					logging.info(traceback.format_exc())
					return None
				try:
					contact_info = self.application.get_profile_contact_info(profile_url)
				except Exception as error:
					logging.exception(f'Error loading contact info: {error}')
					logging.info(traceback.format_exc())
					contact_info = {}
			else:
				try:
					# a sample profile for debugging purposes
					profile = {'lastName': 'SquarePants', 'firstName': 'SpongeBob', 'industryName': 'Professional retard', 'profile_id': f'DEBUG-{random.randint(0,99999)}'}
					contact_info = {'email_address': 'squarepants@bikinibottom.com', 'websites': ['square@pants.bk'], 'twitter': '@pants', 'phone_numbers': ['+001']}
				except Exception as error:
					logging.exception(f'Error loading profile: {error}')
					print(f'⛔️ Error loading profile: {error}')
					return None

			self.store_profile(profile, contact_info)


	def store_profile(self, profile, contact_info):
		def set_diff(dict, full_set):
			'''
			Calculate the difference between the full key set and the provided key set.
			Return the keys that exist in the dictionary, so the missing ones can be
			set to Nonetypes.
			'''
			ignored_keys = {key for key in full_set if key not in dict.keys()}
			return full_set.difference(ignored_keys)

		# the full set of keys a complete profile would have
		profile_keys_full = {
		'firstName', 'lastName', 'profile_id', 'headline', 
		'summary', 'industryName', 'geoCountryName', 'languages'}
		
		contact_keys_full = {'birthdate', 'email_address', 'phone_numbers'}

		# if the profile is lacking keys, replace their values with Nonetypes
		profile_keys = set_diff(profile, profile_keys_full)
		contact_keys = set_diff(contact_info, contact_keys_full)

		# map profile keys to CRM-compatible column names
		column_map = {
		'firstName': 'First name',
		'lastName': 'Last name',
		'profile_id': 'Linkedin profile ID',
		'headline': 'Linkedin headline',
		'summary': 'Linkedin summary',
		'industryName': 'Industry',
		'geoCountryName': 'Location',
		'languages': 'Languages',
		'birthdate': 'Birthday',
		'email_address': 'Email address',
		'phone_numbers': 'Phone number'
		}

		# generate the profile: this is stored later
		profile_dict = {}

		# generate the profile_dict: map API resp. keys to column names, add Nonetypes
		for key in profile_keys_full:
			if key == 'languages' and key in profile_keys:
				# languages: a list of dictionaries with name:value
				try:
					if type(profile['languages']) == list:
						if len(profile['languages']) != 0:
							language_string = ''
							for dict in profile['languages']:
								language_string += dict['name']
								language_string += ', '

							profile['languages'] = language_string[0:-2]
						else:
							profile['languages'] = ''

				except Exception as e:
					profile['languages'] = ''
					logging.exception(f'Error setting language: {e}')
					logging.info(traceback.format_exc())


			if key in profile_keys:
				profile_dict[column_map[key]] = profile[key]
			else:
				profile_dict[column_map[key]] = ''

		# same as above, but for contact keys
		for key in contact_keys_full:
			if key == 'phone_numbers' and key in contact_keys:
				try:
					for val in contact_info['phone_numbers']:
						if len(contact_info['phone_numbers']) > 0:
							numbers = ''
							for dict in contact_info['phone_numbers']:
								numbers += dict['number']
								numbers += f' ({dict["type"]})'
								numbers += ', '

					contact_info['phone_numbers'] = numbers[0:-2]
				except:
					contact_info['phone_numbers'] = ''

			if key in contact_keys:
				profile_dict[column_map[key]] = contact_info[key]
			else:
				profile_dict[column_map[key]] = ''

		logging.info(f'profile_dict generated: {profile_dict}')

		# if this contact is not a duplicate, or we are ignoring duplicates, continue: else, return
		if not self.history.add(profile_dict['Linkedin profile ID'], self.ignore_duplicates):
			#sg.popup('This profile has already been added: avoiding duplicate.', font=('Helvetica', 11), title='Duplicate', keep_on_top=True)
			print(f'⚠️ Duplicate detected ({profile_dict["Linkedin profile ID"]})\n')
			return

		if self.sheet_type == 'csv':
			field_names = profile_dict.keys()
			if not os.path.isfile(self.sheet_path) and self.sheet_type == 'csv':
				with open(self.sheet_path, 'w', newline='') as csv_file:
					csv.DictWriter(csv_file, fieldnames=field_names).writeheader()
					print(f'Created file: {self.sheet_path}')

			with open(self.sheet_path, 'a', newline='') as csv_file:
				csv.DictWriter(csv_file, fieldnames=field_names).writerow(profile_dict)

		elif self.sheet_type == 'excel':
			# convert dictionary to a dataframe
			for key, val in profile_dict.items():
				profile_dict[key] = [val]
			try:
				df = pd.DataFrame(profile_dict, columns=column_map.values())
			except Exception as error:
				logging.exception(f'Exception creating df: {error}')
				logging.info(traceback.format_exc())

			# store (file exists)
			if os.path.isfile(self.sheet_path):
				try:
					book = load_workbook(self.sheet_path)
					with pd.ExcelWriter(self.sheet_path, engine='openpyxl') as writer:
						writer.book = book
						writer.sheets = {ws.title: ws for ws in book.worksheets}
						for sheetname in writer.sheets:
							df.to_excel(
							writer, sheet_name=sheetname, 
							startrow=writer.sheets[sheetname].max_row, 
							index = False, header= False
							)

				except Exception as e:
					logging.exception(f'Error storing profile in file: {e}')
					logging.info(traceback.format_exc())
			else:
				try:
					with pd.ExcelWriter(self.sheet_path, engine='openpyxl') as writer:
						df.to_excel(writer, sheet_name='Sheet1', index=False, header=True)

				except Exception as e:
					logging.exception(f'Error storing first profile in file: {e}')
					logging.info(traceback.format_exc())

		print(f'✅ Stored profile {profile_dict["Linkedin profile ID"]} to {self.sheet_path}\n')
		logging.info(f'Stored profile {profile_dict["Linkedin profile ID"]} to {self.sheet_path}')

		self.parsed += 1
		self.total_parsed += 1


if __name__ == '__main__':
	# create session, start log
	session = Session()
	session.start_log()

	# theme
	load_theme = session.load_theme()
	sg.theme(load_theme)

	# load UI
	session.gui.display_signin_screen()
	logging.info('Program started')

	# sign-in eventloop
	while True:
		event, values = session.gui.window.read()

		if event == sg.WIN_CLOSED:
			logging.info('Sign-in window closed')
			break

		if values['debug_mode']:
			session.debug = True
			session.history.hourly_limit = None

		if event == 'theme_switch':
			toggled_theme = 'DarkBlack' if load_theme == 'SystemDefault' else 'SystemDefault'
			inverse_theme = 'DarkBlack' if load_theme == 'SystemDefault' else 'SystemDefault'
			if values['theme_switch']:
				sg.theme(toggled_theme)
				session.save_theme(toggled_theme)
			else:
				sg.theme(inverse_theme)
				session.save_theme(inverse_theme)

			session.gui.window.finalize()

		elif event == 'debug_screen':
			session.gui.display_debug_screen()
			while True:
				event, values = session.gui.secondary_window.read()
				if event == sg.WIN_CLOSED:
					session.gui.secondary_window.close()
					event, values = None, None
					break

				if event == 'Clear log':
					session.clear_log()

				elif event == 'Clear configuration':
					session.clear_config()

				elif event == 'Remove contacts':
					session.remove_contacts()

				elif event == 'Run tests':
					pass

		elif event == 'show_log':
			session.gui.window['output_window'].update(session.load_log())

		elif event == 'Sign in':
			if values['username'] != '' and values['password'] != '' or values['-USERNAME-'] != [] or session.debug:
				if values['-USERNAME-'] != []:
					username = values['-USERNAME-'][0]
					password = session.load_password_from_config(username)
					values['remember'] = False
				else:
					username = values['username']
					password = values['password']

				logging.info(f'Signing in with stored login: {username} ({type(username)}): {password} ({type(password)})')

				print('Signing in...')
				if not session.debug:
					auth_success = session.sign_in(username, password, values['remember'], values['cookies'])
				else:
					logging.info('Authenticated with debug mode enabled')
					session.username = 'debug user'
					session.authenticated = True
					auth_success = True

					session.sheet_type = session.default_sheet_type
					if session.sheet_type == 'csv':
						session.sheet_path = 'linkedin_scrape.csv'
					elif session.sheet_type == 'excel':
						session.sheet_path = 'linkedin_scrape.xlsx'

					session.gui.window.close()
					session.gui.display_main_screen()
					break

				if not auth_success:
					print('Failed to sign in.\n')
				else:
					sg.popup('Signed in successfully!', title='Success', keep_on_top=True)
					session.gui.window.close()

					# request sheet/csv location
					session.gui.display_sheet_screen()
					while session.sheet_path is None:
						event, values = session.gui.window.read()
						if event == 'Use default' or (event == sg.WIN_CLOSED and session.sheet_path is None):
							session.sheet_type = session.default_sheet_type
							if session.sheet_type == 'csv':
								session.sheet_path = 'linkedin_scrape.csv'
							elif session.sheet_type == 'excel':
								session.sheet_path = 'linkedin_scrape.xlsx'

							if event == sg.WIN_CLOSED:
								sg.popup(
								f'No file path defined. Using default path: {session.sheet_path}', 
								title='No path defined', keep_on_top=True)

							break

						if values['sheet_path'] != '':
							if '.csv' in values['sheet_path']:
								session.sheet_path = values['sheet_path']
								session.sheet_type = 'csv'
							elif '.xls' in values['sheet_path']:
								session.sheet_path = values['sheet_path']
								session.sheet_type = 'excel'

					try:
						session.load_sheet_length()
						session.gui.window.close()
						session.gui.display_main_screen()
					except Exception as error:
						logging.exception(error)

					break
			else:
				sg.popup('Please enter your login details!', title='Incorrect login', keep_on_top=True)

	# main eventloop
	try:
		pipeline = queue.Queue(maxsize=10)
		threading_event = threading.Event()

		with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
			while True and session.authenticated:
				event, values = session.gui.window.read()

				if event == sg.WIN_CLOSED:
					logging.info('Main window closed')

					session.history.store()
					session.gui.window.close()

					logging.info('Exiting main event loop gracefully')
					break

				if event == 'Store contact' and (values['profile_url'] != '' or session.debug):
					if not session.debug:
						print(f'⏳ Loading {values["profile_url"]}...')
						values['profile_url'] = values['profile_url'].split('?')[0]
						if values['profile_url'][-1] == '/':
							values['profile_url'] = values['profile_url'][0:-1]

						if '/' in values['profile_url']:
							profile = values['profile_url'].split('/')[-1]
						else:
							profile = values['profile_url']

						logging.info(f'\nParsing profile {profile}')
					else:
						profile = None
						print('⏳ Parsing sample debug profile...')

					validity_status, time_until_next = session.history.check_validity()
					if validity_status:
						logging.info(f'Profile {profile} put into pipeline...')
						pipeline.put(profile)

						executor.submit(session.linkedin_api_call, pipeline, threading_event)
						threading_event.set()

						# clear input
						session.gui.window['profile_url'].update('')
						session.gui.window['parsed'].update(f'{session.parsed} {"contact" if session.parsed == 1 else "contacts"} stored (this session)\t')
						session.gui.window['total_parsed'].update(f'Contacts in file: {session.total_parsed}\t')

					else:
						sg.popup(f'API call limit reached. Try again in {time_until_next}.', font=('Helvetica', 11), title='Limit reached', keep_on_top=True)
						logging.info(f'API call limit reached: time until next call {time_until_next}. Limit: {session.history.hourly_limit} calls per hour.')

	except Exception as error:
		logging.exception(error)
		session.history.store()
		session.gui.window.close()
